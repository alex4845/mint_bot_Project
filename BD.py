import os
import os.path
import psycopg2
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

async def telegramm_base(state):
    async with state.proxy() as data:
        conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                                password='20rasputin23', database='rasputin_base.db')
        cursor = conn.cursor()
        tims = datetime.now()
        cursor.execute("""CREATE TABLE IF NOT EXISTS list_1
               (number SERIAL PRIMARY KEY, name VARCHAR, insta VARCHAR, username VARCHAR, user_id NUMERIC(20),
                sex VARCHAR, qr BYTEA, time_m TIMESTAMP, manager VARCHAR)""")
        cursor.execute('INSERT INTO list_1 (name, insta, username, user_id, sex, qr, time_m, manager)'
                       ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
                       (data['name'], data['insta'], data['username'], data['user_id'], data['sex'], data['qr'], tims, "--"))
        conn.commit()
        conn.close()

async def get_info():
    list = []
    conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                            password='20rasputin23', database='rasputin_base.db')
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT (*) FROM list_1")
    r = cursor.fetchone()
    list.append(r)
    cursor.execute("SELECT COUNT (*) FROM list_1 WHERE sex = 'М'")
    r1 = cursor.fetchone()
    list.append(r1)
    cursor.execute("SELECT COUNT (*) FROM list_1 WHERE sex = 'Ж'")
    r2 = cursor.fetchone()
    list.append(r2)
    conn.close()
    return list

async def get_user(a):
    conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                            password='20rasputin23', database='rasputin_base.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM list_1 WHERE user_id = %s", (a,))
    res = cursor.fetchone()
    conn.close()
    return res

async def del_user(a):
    conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                            password='20rasputin23', database='rasputin_base.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM list_1 WHERE user_id = %s", (a,))
    res = cursor.fetchone()
    if res is not None:
        record_id = res[0]
        cursor.execute("DELETE FROM list_1 WHERE number = %s", (record_id,))
        conn.commit()
        b = f"Пользователь {res[1]} удален"
    else:
        b = "Пользователь не найден в базе данных"
    conn.close()
    return b

async def get_info_act():
    conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                            password='20rasputin23', database='rasputin_base.db')
    cursor = conn.cursor()
    cursor.execute("SELECT number, name, username, user_id, sex, sur, w_c, w_s, vin, manager FROM list_2")
    res = cursor.fetchall()
    conn.close()
    return res

async def interval():
    l = []
    conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                            password='20rasputin23', database='rasputin_base.db')
    cursor = conn.cursor()
    t_n = datetime.now()
    t_n1 = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().replace(hour=6, minute=0, second=0) #конец работы
    current_time1 = datetime.now().replace(hour=22, minute=30, second=0) #начало работы
    if t_n > current_time and t_n < current_time1:
        cursor.execute('SELECT * FROM list_2')
        data = cursor.fetchall()
        if data:
            data_m = await all_managers()
            folder_name = "reports"
            filename = f'Отчет_{t_n1}.xlsx'
            folder_path = os.path.join(os.getcwd(), folder_name)
            file_path = os.path.join(folder_path, filename)
            workbook = Workbook()
            sheet = workbook.active
            headers = ['Номер', '__Логин__', '__Username__', '____ID____', 'Пол', 'Время получения последнего угощения',
                       'Статус угощения', 'Виски-кола', 'Виски-сок', 'Вино', 'Менеджер']
            for col_index, header in enumerate(headers, start=1):
                column_letter = get_column_letter(col_index)
                sheet[f"{column_letter}1"] = header

            for row_index, row_data in enumerate(data, start=3):
                for col_index, col_data in enumerate(row_data, start=1):
                    sheet.cell(row=row_index, column=col_index, value=col_data)

            sheet['M3'] = str(data_m)[1:-1]

            for col_index, header in enumerate(headers, start=1):
                    column_letter = get_column_letter(col_index)
                    column_dimensions = sheet.column_dimensions[column_letter]
                    column_dimensions.width = len(str(header)) + 2

            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            workbook.save(file_path)

            cursor.execute("TRUNCATE TABLE list_2 RESTART IDENTITY")
            conn.commit()
            conn.close()
            return "fin"
        else:
            conn.close()
    else:
        cursor.execute("SELECT time, number, user_id, w_c, w_s, vin FROM list_2 WHERE sur = %s", ("---",))
        res = cursor.fetchall()
        if res:
            for i in res:
                if int(i[3]) + int(i[4]) + int(i[5]) >= 5: #Максимум угощений
                    continue
                time_1 = datetime.strptime(str(i[0]), "%Y-%m-%d %H:%M:%S.%f")
                diff_minutes = t_n - time_1
                minutes = diff_minutes.total_seconds() // 60
                if minutes > 5: #через сколько минут назначать угощение
                    cursor.execute("UPDATE list_2 SET sur = %s WHERE number = %s", ('+', i[1]))
                    l.append(i[2])
                conn.commit()
            conn.close()
            return l

async def get_user_act(a):
    conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                            password='20rasputin23', database='rasputin_base.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM list_2 WHERE user_id = %s", (a,))
    res = cursor.fetchone()
    conn.close()
    return res

async def write_manager(a, u):
    conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                            password='20rasputin23', database='rasputin_base.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM list_1 WHERE user_id = %s", (u,))
    res = cursor.fetchone()
    time_1 = datetime.strptime(str(res[-2]), "%Y-%m-%d %H:%M:%S.%f")
    time_n = datetime.now()
    diff_minutes = time_n - time_1
    minutes = diff_minutes.total_seconds() // 60
    if res[-1] == "--" or minutes > 1440: #через сколько минут можно выбрать нового менеджера
        cursor.execute("UPDATE list_1 SET manager = %s WHERE user_id = %s", (a, u))
        cursor.execute("UPDATE list_1 SET time_m = %s WHERE user_id = %s", (time_n, u))
        conn.commit()
        conn.close()
        return f"Ваш менеджер {a}. Спасибо."
    else:
        return f"Вы уже выбрали менеджера на сегодня. Это {res[-1]}."

async def all_managers():
    conn = psycopg2.connect(host='141.8.199.12', port=5432, user='postgres',
                            password='20rasputin23', database='rasputin_base.db')
    cursor = conn.cursor()
    b = ['Настя', 'Феми', 'Суворов', 'Марвин', 'Миша', 'Никто']
    rr = []
    for i in b:
        cursor.execute(f"SELECT COUNT(*) FROM list_1 WHERE manager = '{i}'")
        r = cursor.fetchone()
        rr.append(r[0])
    rrr = dict(zip(b, rr))
    return rrr



